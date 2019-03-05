from openpyxl.styles import colors
from openpyxl.styles import Alignment, Border, Color, Font, PatternFill, Protection, Side
from openpyxl.utils import get_column_letter, column_index_from_string
import Workout
import datetime


class Style:

  class Settings:
  
    WHITE='00FFFFFF'
    LIGHTBLACK='00282828'
    DARKGREY='00505050'
    DARKRED='00600000'
    
    ALIGNMENT = Alignment(
        wrap_text=True, horizontal="center", vertical="center"
    )

  @staticmethod
  def generate_header(row: int, col: int, length: int, currentSheet: object, heading: str = 'Header', value: str = 'Item') -> object:
              # Add horizontal header
              # [ Day 1 ]
              currentSheet.merge_cells(
                  start_row=row, end_row=row,
                  start_column=col, end_column=col+length
              )

              currentCell = currentSheet.cell(
                  row=row, column=col, value=f"{heading} {value}"
              )

              return currentCell


  @staticmethod
  def generate_block(row: int, col: int, currentSheet: object, value: str = 'Item') -> object:
              # Add a single cell header
              # [ Block ]

              currentCell = currentSheet.cell(
                  row=row, column=col, value=f"{value}"
              )

              Style.set_style(
                  currentSheet, currentCell, col,
                  fgColor=colors.WHITE, bgColor=Style.Settings.DARKRED,
                  size=18, width=20, font='Helvetica', bold=True
              )

              currentCell.alignment = Style.Settings.ALIGNMENT

              return currentCell


  @staticmethod
  def generate_sheet_banner(currentSheet: object, value: str = 'Item') -> None:
              # Get last column of spreadsheet for full banner
              max_col = currentSheet.max_column

              # Print week banner on first row
              currentSheet.merge_cells(
                  start_row=1, end_row=1,
                  # Calculate total columns in sheet
                  start_column=1, end_column=max_col
              )
              currentCell = currentSheet.cell(
                  row=1, column=1, value=f"{value}"
              )
              Style.set_style(
                  currentSheet, currentCell, 1,
                  fgColor=colors.WHITE, bgColor=Style.Settings.DARKRED,
                  size=28, width=10, font='Helvetica', bold=True
              )
              currentCell.alignment = Style.Settings.ALIGNMENT

              # Print date banner on second row
              date = datetime.datetime.now().strftime("%m/%d/%Y")
              currentSheet.merge_cells(
                  start_row=2, end_row=2,
                  # Calculate total columns in sheet
                  start_column=1, end_column=max_col
              )
              currentCell = currentSheet.cell(
                  row=2, column=1, value=f"Generated by Time to Train on {date}"
              )
              currentCell.hyperlink = "https://github.com/jonschipp/timetotrain"
              currentCell.style = "Hyperlink"
              Style.set_style(
                  currentSheet, currentCell, 1,
                  fgColor=colors.WHITE, bgColor=Style.Settings.LIGHTBLACK,
                  size=12, width=10, font='Helvetica', bold=False
              )
              currentCell.alignment = Style.Settings.ALIGNMENT


  @staticmethod
  def generate_divide(row: int, col: int, length: int, currentSheet: object, heading: str = 'Header', style: str = 'manual') -> object:
              # Create divide with header and input
              # [         ][         ]
              # [ Program ][ <input> ]
              # [         ][         ]
              color = Style.Settings.DARKGREY
              bold = False

              if style == 'formula':
                  color = Style.Settings.DARKRED
                  bold = True

              currentCell = currentSheet.cell(
                  row=row, column=col, value=f"{heading}"
              )

              currentSheet.merge_cells(
                  start_row=row, end_row=row, start_column=col+1, end_column=col+length
              )

              Style.set_style(
                  currentSheet, currentCell, col,
                  fgColor=colors.WHITE, bgColor=color,
                  size=12, width=20, font='Helvetica', bold=bold
              )

              currentCell = currentSheet.cell(
                  row=row, column=col+1
              )

              if style == 'formula':
                  Style.set_style(
                      currentSheet, currentCell, col,
                      fgColor=colors.WHITE, bgColor=color,
                      size=12, width=20, font='Helvetica', bold=False
                  )

              currentCell.alignment = Style.Settings.ALIGNMENT

              return currentCell


  @staticmethod
  def set_style(sheet: object, cell: object, col: int, fgColor: str, bgColor: str, size: int, width: int, font: str, bold: bool = False) -> object:
        # Set style
        font = Font(
            name=font, size=size, bold=bold, color=fgColor
        )
        fill = PatternFill(
            fill_type='solid', fgColor=bgColor,
        )

        sheet.column_dimensions[get_column_letter(col)].width = width
        cell.font = font
        cell.fill = fill
        cell.alignment = Style.Settings.ALIGNMENT
        return cell
